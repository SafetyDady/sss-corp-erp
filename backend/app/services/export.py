"""
SSS Corp ERP — Shared Excel Export Helper
Phase 10: Excel export using openpyxl

Creates styled .xlsx workbooks with company header, column headers, and data rows.
"""

from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Styles ──────────────────────────────────────────────────
_HEADER_FONT = Font(name="Tahoma", size=14, bold=True)
_SUBHEADER_FONT = Font(name="Tahoma", size=10, color="666666")
_COL_HEADER_FONT = Font(name="Tahoma", size=10, bold=True, color="FFFFFF")
_COL_HEADER_FILL = PatternFill(start_color="2A2D35", end_color="2A2D35", fill_type="solid")
_DATA_FONT = Font(name="Tahoma", size=10)
_MONEY_FONT = Font(name="Consolas", size=10)
_THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
_CENTER = Alignment(horizontal="center", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def create_excel_workbook(
    *,
    title: str,
    headers: list[str],
    rows: list[list],
    org_name: Optional[str] = None,
    col_widths: Optional[list[int]] = None,
    money_cols: Optional[list[int]] = None,
) -> BytesIO:
    """
    Create a styled .xlsx workbook with optional company header.

    Args:
        title: Sheet title & document heading
        headers: Column header labels
        rows: 2D list of row data
        org_name: Company name shown at top (optional)
        col_widths: Column widths (optional, auto-fit if not given)
        money_cols: 0-based column indices that contain money (right-align + #,##0.00)

    Returns:
        BytesIO buffer ready for StreamingResponse
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name max 31 chars

    money_cols = set(money_cols or [])
    current_row = 1

    # ── Company header (optional) ───────────────────────────
    if org_name:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
        cell = ws.cell(row=current_row, column=1, value=org_name)
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        current_row += 1

        # Document title
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
        cell = ws.cell(row=current_row, column=1, value=title)
        cell.font = _SUBHEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        current_row += 2  # blank row
    else:
        current_row = 1

    # ── Column headers ──────────────────────────────────────
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = _COL_HEADER_FONT
        cell.fill = _COL_HEADER_FILL
        cell.border = _THIN_BORDER
        cell.alignment = _CENTER
    current_row += 1

    # ── Data rows ───────────────────────────────────────────
    for row_data in rows:
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.border = _THIN_BORDER

            if (col_idx - 1) in money_cols:
                cell.font = _MONEY_FONT
                cell.alignment = _RIGHT
                cell.number_format = '#,##0.00'
            else:
                cell.font = _DATA_FONT
                cell.alignment = _LEFT
        current_row += 1

    # ── Column widths ───────────────────────────────────────
    if col_widths:
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
    else:
        # Auto-fit: max of header length and data (sample first 50 rows)
        for col_idx, header in enumerate(headers, start=1):
            max_len = len(str(header))
            for row_data in rows[:50]:
                if col_idx - 1 < len(row_data):
                    max_len = max(max_len, len(str(row_data[col_idx - 1] or "")))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

    # ── Freeze header row ───────────────────────────────────
    header_row = (4 if org_name else 1)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # ── Write to buffer ─────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
